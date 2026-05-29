#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Plan Estrategico Dashboard.py
=============================

Generador automático del dashboard HTML del Plan Estratégico 2026-2030.

Lee el archivo Excel del modelo (con las 25 hojas individuales de cada
indicador), procesa los datos mensualmente desde enero 2026 hasta diciembre
2030, calcula resultados, % de cumplimiento, semáforos, tendencias propias,
avance acumulado y alertas, y genera un dashboard HTML interactivo, moderno,
responsive y autocontenido, listo para publicar en GitHub Pages.

Este es un único script Python autocontenido. Solo necesita openpyxl:

    pip install openpyxl

Uso
---
    python3 "Plan Estrategico Dashboard.py"
    python3 "Plan Estrategico Dashboard.py" <ruta_excel>
    python3 "Plan Estrategico Dashboard.py" <ruta_excel> <ruta_salida_html>

Por defecto:
    Input:  Tablero_Indicadores_PE.xlsx
    Output: docs/index.html

Archivo auxiliar opcional en la misma carpeta:
    chart_umd.js   ← Chart.js (si existe se embebe; si no, se carga desde CDN)

Estructura interna del script
-----------------------------
    1.  CATÁLOGO DE INDICADORES        (datos del modelo)
    2.  PLANTILLAS HTML / CSS / JS     (paleta KALLA)
    3.  EXTRACCIÓN DEL EXCEL           (lectura openpyxl)
    4.  CÁLCULOS                       (resultados, cumplimiento, etc.)
    5.  VALIDACIÓN
    6.  CONSTRUCCIÓN DE PAYLOAD JSON
    7.  RENDERIZADO DE HTML
    8.  MAIN
"""

import sys
import json
import math
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl. Instale con:  pip install openpyxl")
    sys.exit(1)



# ============================================================================
# 1. CATÁLOGO DE INDICADORES
# ============================================================================

OBJ_GEN = {
    "FINANCIERA": "Mantener un margen financiero adecuado, buscando una distribución de los ingresos (50% gastos, 35% costos y 15% excedentes), mediante la eficiencia operativa, el control integral del riesgo y la optimización continua de la cartera, apoyado en políticas financieras prudentes y competitivas, herramientas de análisis y un equipo técnico fortalecido, con cumplimiento proyectado a diciembre de 2030.",
    "ASOCIADOS": "Aumentar la base social a 8.000 asociados al año 2030, mediante la implementación de estrategias de fidelización y campañas de atracción segmentadas, apoyado en una estructura comercial fortalecida, canales ágiles de comunicación y el uso de inteligencia de datos.",
    "PROCESOS INTERNOS": "Automatizar al menos el 70% de las tareas susceptibles de automatización asociadas a los procesos operativos y de servicio al asociado al finalizar el año 2030, apoyándose en tecnología, analítica, seguridad de la información y capacidades digitales organizacionales.",
    "APRENDIZAJE Y CRECIMIENTO": "Potenciar la formación financiera y solidaria de los asociados, así como el desarrollo profesional de los empleados, alcanzando al menos el 50% de participación de asociados y el 100% de empleados, mediante programas de formación continua, alianzas estratégicas y campañas de educación cooperativa, con recursos institucionales y seguimiento de indicadores de desarrollo y participación, a diciembre de 2030.",
}

PERSP_INFO = {
    "FINANCIERA":               {"short": "Financiera",   "icon": "💰", "accent": "1F6FB2"},
    "ASOCIADOS":                {"short": "Asociados",    "icon": "🤝", "accent": "00A14B"},
    "PROCESOS INTERNOS":        {"short": "Procesos",     "icon": "⚙️", "accent": "E08A00"},
    "APRENDIZAJE Y CRECIMIENTO":{"short": "Aprendizaje",  "icon": "🎓", "accent": "8E44AD"},
}

INI_ICON = {
    "Aumento de ingresos":"📈",
    "Reducción de mora":"🛡️",
    "Crecimiento y remuneración de depósitos":"🏦",
    "Rentabilidad y fortalecimiento patrimonial":"💎",
    "Crecimiento de la base social":"🌱",
    "Retención y profundización":"💗",
    "Operatividad digital y canales":"📱",
    "Automatización de procesos":"🤖",
    "Transformación digital del asociado":"💻",
    "Formación de asociados":"📚",
    "Formación y desarrollo del personal":"👥",
}

INI_DESC = {
    "Aumento de ingresos": "Aumentar los ingresos mediante la optimización de la cartera, la colocación de crédito en líneas de mayor rendimiento, la generación de comisiones por servicios no financieros y la rentabilidad de las inversiones.",
    "Reducción de mora": "Reducir y controlar el deterioro de cartera, fortaleciendo la gestión preventiva, la recuperación y la normalización de obligaciones.",
    "Crecimiento y remuneración de depósitos": "Fortalecer la captación y remuneración competitiva de depósitos como base de fondeo estable y rentable.",
    "Rentabilidad y fortalecimiento patrimonial": "Asegurar márgenes financieros sostenibles y consolidar el patrimonio mediante el cumplimiento de aportes y la generación de capital institucional.",
    "Crecimiento de la base social": "Impulsar el crecimiento de la base social mediante la vinculación de nuevos asociados, soportado en ofertas diferenciadas que respondan a las necesidades de los segmentos objetivo.",
    "Retención y profundización": "Reducir la desvinculación de asociados e incrementar el uso del portafolio de servicios mediante acciones de retención, seguimiento y profundización comercial.",
    "Operatividad digital y canales": "Optimizar los canales de atención y comunicación con los asociados, asegurando efectividad y satisfacción en cada interacción.",
    "Automatización de procesos": "Automatizar procesos operativos y de servicio al asociado para mejorar eficiencia, calidad y tiempos de respuesta.",
    "Transformación digital del asociado": "Promover la transformación digital de los asociados, incentivando el uso de los canales transaccionales digitales.",
    "Formación de asociados": "Fortalecer la formación financiera y solidaria de los asociados.",
    "Formación y desarrollo del personal": "Fortalecer el desarrollo profesional de los funcionarios y el clima organizacional.",
}

PERSP_ORDER = ["FINANCIERA","ASOCIADOS","PROCESOS INTERNOS","APRENDIZAJE Y CRECIMIENTO"]

# Indicadores
# tipo: 'ratio' | 'growth' | 'diff' | 'avg' | 'count'
# vars: lista de campos de entrada mensual con (codigo, etiqueta)
# metas: lista de 5 valores (2026..2030); None o str = "por definir" / referencia
# tipo_meta: 'pct' | 'bps' | 'ratio' | 'unidades' | 'texto'
IND = [
 # ============ FINANCIERA ============
 dict(p="FINANCIERA", ini="Aumento de ingresos",
   indicador="Crecimiento de cartera en líneas de mayor rendimiento",
   desc="Mide la participación de las líneas de crédito con mayor tasa efectiva y contribución al margen financiero. Por el momento, Rotativo y Tarjeta de Crédito.",
   formula="Saldo de cartera de líneas de mayor rendimiento ÷ Saldo total de cartera",
   tipo="ratio", vars=[("num","Saldo cartera líneas mayor rendimiento ($)"),("den","Saldo total de cartera ($)")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Mensual",
   metas=[0.08,0.10,0.12,0.14,0.16]),
 dict(p="FINANCIERA", ini="Aumento de ingresos",
   indicador="Crecimiento de ingresos operativos",
   desc="Refleja el impacto directo de la colocación sobre los ingresos financieros.",
   formula="(Ingresos de cartera final − Ingresos de cartera inicial) ÷ Ingresos de cartera inicial",
   tipo="growth", vars=[("val","Ingresos de cartera del mes ($)")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Mensual",
   metas=[0.10,0.10,0.10,0.10,0.10]),
 dict(p="FINANCIERA", ini="Aumento de ingresos",
   indicador="Crecimiento de ingresos por comisiones",
   desc="Mide el incremento de ingresos derivados del uso de servicios no financieros (convenios).",
   formula="(Ingresos por comisiones final − Ingresos por comisiones inicial) ÷ Ingresos por comisiones inicial",
   tipo="growth", vars=[("val","Ingresos por comisiones del mes ($)")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Mensual",
   metas=[0.035,0.035,0.035,0.035,0.035]),
 dict(p="FINANCIERA", ini="Aumento de ingresos",
   indicador="Rentabilidad Fondo de Liquidez",
   desc="Mide la capacidad de generar rentabilidad por encima del mercado.",
   formula="Tasa promedio ponderada de rentabilidad del fondo de liquidez − IBR",
   tipo="diff", vars=[("a","Tasa promedio ponderada Fondo de Liquidez (%)"),("b","IBR (%)")],
   unidad="Puntos básicos", tipo_meta="bps", tend="Crecer", frec="Mensual",
   metas=[-200,-200,-200,-200,-200]),
 dict(p="FINANCIERA", ini="Reducción de mora",
   indicador="Índice de mora",
   desc="Mide el nivel de morosidad de la cartera por crédito, sin aplicación de arrastres o riesgo.",
   formula="Saldo de cartera ≥ B ÷ Saldo total de cartera",
   tipo="ratio", vars=[("num","Saldo de cartera en mora ≥ B ($)"),("den","Saldo total de cartera ($)")],
   unidad="%", tipo_meta="pct", tend="Reducir", frec="Mensual",
   metas=[0.05,0.05,0.05,0.05,0.05]),
 dict(p="FINANCIERA", ini="Reducción de mora",
   indicador="Tasa de recuperación",
   desc="Mide la proporción de cartera recuperada frente a la deteriorada por rodamiento (matriz de transición).",
   formula="Saldo de cartera recuperada ÷ Saldo de cartera deteriorada",
   tipo="ratio", vars=[("num","Saldo cartera recuperada ($)"),("den","Saldo cartera deteriorada ($)")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Mensual",
   metas=[0.30,0.30,0.30,0.30,0.30]),
 dict(p="FINANCIERA", ini="Reducción de mora",
   indicador="Tasa de normalización de cartera",
   desc="Mide la efectividad de acuerdos de pago y reestructuraciones.",
   formula="Saldo de normalizaciones al día ÷ Saldo total de normalizaciones",
   tipo="ratio", vars=[("num","Saldo normalizaciones al día ($)"),("den","Saldo total normalizaciones ($)")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Mensual",
   metas=[0.80,0.80,0.80,0.80,0.80]),
 dict(p="FINANCIERA", ini="Crecimiento y remuneración de depósitos",
   indicador="Tasa de crecimiento de ahorros",
   desc="Mide el crecimiento de los depósitos entre periodos.",
   formula="(Saldo final depósitos − Saldo inicial depósitos) ÷ Saldo inicial depósitos",
   tipo="growth", vars=[("val","Saldo total de depósitos del mes ($)")],
   unidad="%", tipo_meta="texto", tend="Crecer", frec="Mensual",
   metas=["≥ IPC","≥ IPC","≥ IPC","≥ IPC","≥ IPC"]),
 dict(p="FINANCIERA", ini="Crecimiento y remuneración de depósitos",
   indicador="Rentabilidad real de los depósitos",
   desc="Asegura que la tasa de rendimiento de los depósitos supere la inflación (poder adquisitivo del ahorro).",
   formula="Tasa promedio ponderada ahorros − IPC",
   tipo="diff", vars=[("a","Tasa promedio ponderada ahorros (%)"),("b","IPC (%)")],
   unidad="% vs IPC", tipo_meta="pct", tend="Mantener", frec="Mensual",
   metas=[0,0,0,0,0]),
 dict(p="FINANCIERA", ini="Rentabilidad y fortalecimiento patrimonial",
   indicador="Suficiencia del Margen",
   desc="Evalúa la capacidad del margen financiero bruto para cubrir los gastos operativos.",
   formula="Margen financiero bruto ÷ Gastos operativos",
   tipo="ratio", vars=[("num","Margen financiero bruto ($)"),("den","Gastos operativos ($)")],
   unidad="Veces (x)", tipo_meta="ratio", tend="Mantener", frec="Mensual",
   metas=[1.2,1.2,1.2,1.2,1.2]),
 dict(p="FINANCIERA", ini="Rentabilidad y fortalecimiento patrimonial",
   indicador="Cumplimiento de pago de aportes sociales",
   desc="Mide el grado de cumplimiento del recaudo mensual de aportes frente al esperado.",
   formula="Recaudo real de aportes ÷ Recaudo esperado de aportes",
   tipo="ratio", vars=[("num","Recaudo real de aportes ($)"),("den","Recaudo esperado de aportes ($)")],
   unidad="%", tipo_meta="pct", tend="Mantener", frec="Mensual",
   metas=[0.95,0.95,0.95,0.95,0.95]),
 dict(p="FINANCIERA", ini="Rentabilidad y fortalecimiento patrimonial",
   indicador="Capital institucional",
   desc="Mide la proporción del patrimonio permanente frente al total de activos.",
   formula="(Aportes amortizados + Reserva protección aportes + Excedente mes) ÷ Activo",
   tipo="ratio", vars=[("n1","Aportes amortizados ($)"),("n2","Reserva protección aportes ($)"),("n3","Excedente del mes ($)"),("den","Activo total ($)")],
   unidad="%", tipo_meta="pct", tend="Mantener", frec="Mensual",
   metas=[0.15,0.15,0.15,0.15,0.15]),
 dict(p="FINANCIERA", ini="Rentabilidad y fortalecimiento patrimonial",
   indicador="ROE",
   desc="Rentabilidad sobre el patrimonio.",
   formula="Excedente ÷ Patrimonio",
   tipo="ratio", vars=[("num","Excedente ($)"),("den","Patrimonio ($)")],
   unidad="%", tipo_meta="texto", tend="Mantener", frec="Mensual",
   metas=["> IPC","> IPC","> IPC","> IPC","> IPC"]),
 # ============ ASOCIADOS ============
 dict(p="ASOCIADOS", ini="Crecimiento de la base social",
   indicador="Crecimiento de la base social",
   desc="Evalúa el incremento de asociados y el cumplimiento de la meta de base social.",
   formula="((Asociados periodo actual − periodo anterior) ÷ periodo anterior); meta = total asociados",
   tipo="count", vars=[("val","Total de asociados al cierre del mes")],
   unidad="N° asociados", tipo_meta="unidades", tend="Crecer", frec="Semestral",
   metas=[5846,6384,6923,7461,8000]),
 dict(p="ASOCIADOS", ini="Retención y profundización",
   indicador="Tasa de retención de asociados anual",
   desc="Mide la capacidad de la cooperativa para mantener su base social activa.",
   formula="(Total asociados final − asociados nuevos) ÷ Total asociados inicial",
   tipo="ratio", vars=[("n1","Total asociados al final del periodo"),("n2","Asociados nuevos del periodo"),("den","Total asociados al inicio del periodo")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Anual",
   metas=[0.98,0.98,0.98,0.98,0.98]),
 dict(p="ASOCIADOS", ini="Retención y profundización",
   indicador="Índice de satisfacción del asociado",
   desc="Evalúa la percepción del servicio para identificar oportunidades de mejora.",
   formula="Promedio ponderado de encuestas de satisfacción",
   tipo="avg", vars=[("val","Resultado de encuesta de satisfacción (escala 0–1)")],
   unidad="Puntos (escala)", tipo_meta="pct", tend="Crecer", frec="Mensual",
   metas=[0.95,0.95,0.95,0.95,0.95]),
 dict(p="ASOCIADOS", ini="Retención y profundización",
   indicador="Promedio de productos por asociado",
   desc="Mide el nivel de profundización comercial en la base social.",
   formula="Total de productos activos ÷ Total de asociados activos",
   tipo="ratio", vars=[("num","Total productos activos"),("den","Total asociados activos")],
   unidad="Productos", tipo_meta="ratio", tend="Crecer", frec="Mensual",
   metas=[3,3,3,3,3]),
 dict(p="ASOCIADOS", ini="Operatividad digital y canales",
   indicador="Tasa de contacto efectivo por canal",
   desc="Evalúa la efectividad de los canales en la interacción con los asociados.",
   formula="Contactos efectivos ÷ Total contactos realizados",
   tipo="ratio", vars=[("num","Contactos efectivos"),("den","Total contactos realizados")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Mensual",
   metas=[None,None,None,None,None]),
 dict(p="ASOCIADOS", ini="Operatividad digital y canales",
   indicador="Índice de satisfacción por canal",
   desc="Mide la percepción del asociado frente a cada canal de atención.",
   formula="Promedio de calificación obtenida en encuestas por canal",
   tipo="avg", vars=[("val","Calificación promedio encuesta por canal (escala 0–1)")],
   unidad="Puntos (escala)", tipo_meta="pct", tend="Crecer", frec="Trimestral",
   metas=[None,None,None,None,None]),
 # ============ PROCESOS ============
 dict(p="PROCESOS INTERNOS", ini="Automatización de procesos",
   indicador="Nivel de automatización de procesos",
   desc="Mide el porcentaje de procesos operativos y de servicio que han sido automatizados.",
   formula="Tareas automatizadas ÷ Total procesos identificados",
   tipo="ratio", vars=[("num","Tareas/procesos automatizados"),("den","Total procesos identificados")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Semestral",
   metas=[0.30,0.45,0.55,0.65,0.70]),
 dict(p="PROCESOS INTERNOS", ini="Transformación digital del asociado",
   indicador="Uso y adopción de canales transaccionales digitales",
   desc="Mide el uso de canales digitales (Coop Digital y App) por asociados menores de 70 años.",
   formula="Transacciones digitales (<70 años) ÷ Total transacciones (<70 años)",
   tipo="ratio", vars=[("num","Transacciones digitales (<70 años)"),("den","Total transacciones (<70 años)")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Trimestral",
   metas=[0.80,0.80,0.80,0.80,0.80]),
 # ============ APRENDIZAJE ============
 dict(p="APRENDIZAJE Y CRECIMIENTO", ini="Formación de asociados",
   indicador="Formación de asociados por temáticas (financiera - solidaria)",
   desc="Mide el porcentaje de asociados que participan en programas de educación financiera y cooperativa.",
   formula="Asociados capacitados ÷ Total de asociados",
   tipo="ratio", vars=[("num","Asociados capacitados (acumulado año)"),("den","Total de asociados")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Trimestral",
   metas=[0.40,0.40,0.40,0.40,0.40]),
 dict(p="APRENDIZAJE Y CRECIMIENTO", ini="Formación de asociados",
   indicador="Formación en economía solidaria (20 hrs mínimo)",
   desc="Mide el porcentaje de asociados que han cumplido con la formación básica de economía solidaria.",
   formula="Asociados con horas mínimas de formación ÷ Total de asociados",
   tipo="ratio", vars=[("num","Asociados que cumplen 20h"),("den","Total de asociados")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Trimestral",
   metas=[0.40,0.40,0.40,0.40,0.40]),
 dict(p="APRENDIZAJE Y CRECIMIENTO", ini="Formación y desarrollo del personal",
   indicador="Cumplimiento del plan de formación de personal",
   desc="Mide el grado de ejecución del plan de capacitación definido.",
   formula="Actividades de formación ejecutadas ÷ Actividades planificadas",
   tipo="ratio", vars=[("num","Actividades de formación ejecutadas"),("den","Actividades planificadas")],
   unidad="%", tipo_meta="pct", tend="Crecer", frec="Semestral",
   metas=[1.00,1.00,1.00,1.00,1.00]),
 dict(p="APRENDIZAJE Y CRECIMIENTO", ini="Formación y desarrollo del personal",
   indicador="Clima organizacional y compromiso del personal",
   desc="Mide la percepción de los colaboradores sobre el ambiente laboral.",
   formula="Promedio de resultados de la encuesta de clima organizacional",
   tipo="avg", vars=[("val","Resultado encuesta clima (escala 0–1)")],
   unidad="Puntos (escala)", tipo_meta="pct", tend="Mantener", frec="Anual",
   metas=[0.85,0.85,0.85,0.85,0.85]),
]

# Sheet code for each indicator
for i,d in enumerate(IND,1):
    d["code"] = f"IND-{i:02d}"


# ============================================================================
# 2. PLANTILLAS HTML / CSS / JAVASCRIPT
# ============================================================================

import json


# ------------------------------------------------------------
# Paleta inspirada en la rueda KALLA (verdes, teales, azules)
# ------------------------------------------------------------
PALETTE = {
    # Marca
    'primary':       '#0876B2',   # azul KALLA
    'primary_dark':  '#075A8A',
    'primary_light': '#2989D0',
    'teal':          '#02A2C6',
    'teal_dark':     '#007C95',
    'green':         '#23A968',
    'green_light':   '#45BA76',
    # Semáforo
    'success':       '#23A968',
    'warning':       '#F0A030',
    'danger':        '#E04B4B',
    'neutral':       '#94A6B8',
    # Superficies
    'bg':            '#F4F8FB',
    'surface':       '#FFFFFF',
    'surface_alt':   '#EEF4F8',
    'border':        '#DCE6EE',
    # Texto
    'text_dark':     '#0E2A40',
    'text_med':      '#3F5A73',
    'text_lt':       '#7990A4',
}

# Color por perspectiva (dentro de la familia KALLA)
PERSP_COLOR = {
    'FINANCIERA':                '#0876B2',  # azul
    'ASOCIADOS':                 '#23A968',  # verde
    'PROCESOS INTERNOS':         '#02A2C6',  # teal
    'APRENDIZAJE Y CRECIMIENTO': '#1A4D8C',  # azul profundo
}


# ============================================================
# CSS
# ============================================================

CSS = r"""
:root {
  --kalla-primary:       #0876B2;
  --kalla-primary-dark:  #075A8A;
  --kalla-primary-light: #2989D0;
  --kalla-teal:          #02A2C6;
  --kalla-teal-dark:     #007C95;
  --kalla-green:         #23A968;
  --kalla-green-light:   #45BA76;
  --success: #23A968;
  --warning: #F0A030;
  --danger:  #E04B4B;
  --neutral: #94A6B8;
  --bg:        #F4F8FB;
  --surface:   #FFFFFF;
  --surface-2: #EEF4F8;
  --border:    #DCE6EE;
  --text-dark: #0E2A40;
  --text-med:  #3F5A73;
  --text-lt:   #7990A4;
  --shadow-sm: 0 1px 2px rgba(14, 42, 64, 0.04);
  --shadow:    0 4px 12px rgba(14, 42, 64, 0.06);
  --shadow-lg: 0 8px 28px rgba(14, 42, 64, 0.10);
  --radius:    12px;
  --radius-sm: 8px;
  --radius-lg: 18px;
}

* { box-sizing: border-box; }
html, body { margin: 0; padding: 0; }
body {
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
  background: var(--bg);
  color: var(--text-dark);
  font-size: 14px;
  line-height: 1.5;
  -webkit-font-smoothing: antialiased;
}

/* ─── Header ─────────────────────────────────────────────── */
.header {
  background: linear-gradient(135deg, var(--kalla-primary) 0%, var(--kalla-teal) 100%);
  color: white;
  padding: 28px 32px;
  box-shadow: var(--shadow);
}
.header-inner {
  max-width: 1400px; margin: 0 auto;
  display: flex; align-items: center; justify-content: space-between;
  flex-wrap: wrap; gap: 20px;
}
.header h1 {
  font-size: 1.6rem; margin: 0; font-weight: 700; letter-spacing: -0.01em;
}
.header .sub { font-size: 0.92rem; opacity: 0.92; margin-top: 4px; }
.header .meta {
  font-size: 0.82rem; opacity: 0.86; text-align: right;
}
.header .meta strong { font-weight: 600; }
.brand-dot {
  display: inline-block; width: 10px; height: 10px; border-radius: 50%;
  background: var(--kalla-green-light); margin-right: 8px;
  box-shadow: 0 0 0 3px rgba(255,255,255,0.20);
}

/* ─── Container ──────────────────────────────────────────── */
.container {
  max-width: 1400px; margin: 0 auto; padding: 24px 32px;
}

/* ─── KPI cards ──────────────────────────────────────────── */
.kpi-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
  gap: 14px; margin-bottom: 22px;
}
.kpi {
  background: var(--surface); border: 1px solid var(--border);
  border-radius: var(--radius); padding: 18px 20px;
  display: flex; align-items: center; gap: 14px;
  box-shadow: var(--shadow-sm); transition: transform .15s, box-shadow .15s;
}
.kpi:hover { transform: translateY(-1px); box-shadow: var(--shadow); }
.kpi-icon {
  width: 44px; height: 44px; border-radius: 12px;
  display: flex; align-items: center; justify-content: center;
  font-size: 1.4rem; flex-shrink: 0;
}
.kpi-icon.total    { background: #E6F0F8; color: var(--kalla-primary); }
.kpi-icon.cumple   { background: #E2F5EC; color: var(--success); }
.kpi-icon.riesgo   { background: #FCF0DF; color: var(--warning); }
.kpi-icon.critico  { background: #FBE6E6; color: var(--danger); }
.kpi-icon.pendiente{ background: #EEF3F7; color: var(--neutral); }
.kpi-value { font-size: 1.6rem; font-weight: 700; line-height: 1; color: var(--text-dark); }
.kpi-label { font-size: 0.78rem; color: var(--text-med); margin-top: 4px; text-transform: uppercase; letter-spacing: 0.04em; }

/* ─── Filters ────────────────────────────────────────────── */
.filters {
  background: var(--surface); border: 1px solid var(--border);
  border-radius: var(--radius); padding: 14px 18px;
  display: flex; flex-wrap: wrap; gap: 18px; align-items: center;
  margin-bottom: 22px; box-shadow: var(--shadow-sm);
}
.filter-group {
  display: flex; align-items: center; gap: 8px; flex-wrap: wrap;
}
.filter-label {
  font-size: 0.78rem; font-weight: 600; color: var(--text-med);
  text-transform: uppercase; letter-spacing: 0.04em; margin-right: 4px;
}
.chip {
  display: inline-flex; align-items: center; gap: 6px;
  padding: 7px 13px; border-radius: 999px;
  background: var(--surface-2); color: var(--text-med);
  font-size: 0.85rem; font-weight: 500; cursor: pointer;
  border: 1px solid transparent; transition: all .15s;
  user-select: none;
}
.chip:hover { background: #E2EDF4; color: var(--text-dark); }
.chip.active {
  background: var(--kalla-primary); color: white;
  box-shadow: 0 2px 6px rgba(8, 118, 178, 0.30);
}
.chip.active.persp-asociados   { background: var(--kalla-green); box-shadow: 0 2px 6px rgba(35,169,104,0.30); }
.chip.active.persp-procesos    { background: var(--kalla-teal);  box-shadow: 0 2px 6px rgba(2,162,198,0.30); }
.chip.active.persp-aprendizaje { background: #1A4D8C;            box-shadow: 0 2px 6px rgba(26,77,140,0.30); }
.chip.year.active { background: var(--text-dark); }

/* ─── Perspective sections ──────────────────────────────── */
.perspective {
  margin-bottom: 32px;
}
.persp-header {
  display: flex; align-items: center; gap: 12px;
  padding: 14px 20px; border-radius: var(--radius);
  color: white; margin-bottom: 14px;
  box-shadow: var(--shadow-sm);
}
.persp-header .icon { font-size: 1.4rem; }
.persp-header .title { font-weight: 700; font-size: 1.05rem; flex: 1; }
.persp-header .count {
  background: rgba(255,255,255,0.2); padding: 4px 12px; border-radius: 999px;
  font-size: 0.8rem; font-weight: 600;
}
.persp-objetivo {
  background: var(--surface); border: 1px solid var(--border);
  border-left: 4px solid currentColor;
  padding: 12px 18px; border-radius: var(--radius-sm);
  margin-bottom: 16px; font-size: 0.88rem; color: var(--text-med);
  line-height: 1.55;
}
.persp-objetivo strong { color: var(--text-dark); font-weight: 600; }

/* ─── Indicator card ─────────────────────────────────────── */
.ind-grid {
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(360px, 1fr));
  gap: 16px;
}
.ind-card {
  background: var(--surface); border: 1px solid var(--border);
  border-radius: var(--radius); padding: 18px 20px;
  box-shadow: var(--shadow-sm);
  display: flex; flex-direction: column; gap: 12px;
  transition: transform .15s, box-shadow .15s, border-color .15s;
  cursor: pointer; position: relative; overflow: hidden;
}
.ind-card:hover { transform: translateY(-2px); box-shadow: var(--shadow); border-color: #C1D3E0; }
.ind-card .accent-bar {
  position: absolute; left: 0; top: 0; bottom: 0; width: 4px;
}
.ind-head { display: flex; justify-content: space-between; gap: 12px; align-items: flex-start; }
.ind-code { font-size: 0.72rem; color: var(--text-lt); font-weight: 600; letter-spacing: 0.06em; }
.ind-name { font-size: 0.98rem; font-weight: 700; color: var(--text-dark); margin: 3px 0 0; line-height: 1.35; }
.ind-ini { font-size: 0.75rem; color: var(--text-med); margin-top: 4px; }

.status-pill {
  display: inline-flex; align-items: center; gap: 5px;
  padding: 5px 10px; border-radius: 999px;
  font-size: 0.72rem; font-weight: 700;
  text-transform: uppercase; letter-spacing: 0.04em;
  white-space: nowrap;
}
.status-pill.cumple    { background: #E2F5EC; color: var(--success); }
.status-pill.riesgo    { background: #FCF0DF; color: var(--warning); }
.status-pill.critico   { background: #FBE6E6; color: var(--danger); }
.status-pill.pendiente { background: #EEF3F7; color: var(--neutral); }

.ind-metrics {
  display: grid; grid-template-columns: 1fr 1fr 1fr;
  gap: 8px; padding: 10px; background: var(--surface-2);
  border-radius: var(--radius-sm);
}
.metric { text-align: center; }
.metric .label { font-size: 0.66rem; color: var(--text-lt); text-transform: uppercase; letter-spacing: 0.05em; }
.metric .value { font-size: 0.95rem; font-weight: 700; color: var(--text-dark); margin-top: 2px; }
.metric .value.cumple    { color: var(--success); }
.metric .value.riesgo    { color: var(--warning); }
.metric .value.critico   { color: var(--danger); }
.metric .value.pendiente { color: var(--neutral); }

.chart-wrap { position: relative; height: 110px; }
.ind-foot {
  display: flex; justify-content: space-between; gap: 8px;
  font-size: 0.74rem; color: var(--text-med);
  border-top: 1px solid var(--border); padding-top: 10px;
}
.foot-cell { display: flex; flex-direction: column; gap: 1px; }
.foot-cell .l { font-size: 0.66rem; color: var(--text-lt); text-transform: uppercase; letter-spacing: 0.04em; }
.foot-cell .v { font-weight: 600; color: var(--text-dark); }
.trend-up   { color: var(--success); }
.trend-down { color: var(--danger); }
.trend-flat { color: var(--text-lt); }

.alert {
  margin-top: -4px;
  padding: 8px 12px; border-radius: var(--radius-sm);
  font-size: 0.78rem; font-weight: 500;
  display: flex; align-items: center; gap: 8px;
}
.alert.ok       { background: #E2F5EC; color: var(--success); }
.alert.warn     { background: #FCF0DF; color: #B07000; }
.alert.critical { background: #FBE6E6; color: #A03030; }
.alert.pending  { background: var(--surface-2); color: var(--text-med); }

/* ─── Modal ──────────────────────────────────────────────── */
.modal-overlay {
  position: fixed; inset: 0; background: rgba(14, 42, 64, 0.55);
  z-index: 100; display: none; align-items: flex-start; justify-content: center;
  padding: 40px 20px; overflow-y: auto;
  animation: fadeIn .15s;
}
.modal-overlay.active { display: flex; }
@keyframes fadeIn { from { opacity: 0; } to { opacity: 1; } }

.modal {
  background: var(--surface); border-radius: var(--radius-lg);
  max-width: 1100px; width: 100%;
  box-shadow: var(--shadow-lg);
  overflow: hidden;
}
.modal-head {
  padding: 24px 28px; border-bottom: 1px solid var(--border);
  display: flex; gap: 20px; justify-content: space-between; align-items: flex-start;
}
.modal-head .titles { flex: 1; min-width: 0; }
.modal-head .code { font-size: 0.78rem; color: var(--text-lt); font-weight: 600; letter-spacing: 0.06em; }
.modal-head .name { font-size: 1.3rem; font-weight: 700; color: var(--text-dark); margin: 4px 0; }
.modal-head .persp-tag { font-size: 0.82rem; color: var(--text-med); }
.modal-close {
  background: var(--surface-2); border: none; width: 36px; height: 36px;
  border-radius: 10px; cursor: pointer; font-size: 1.4rem; color: var(--text-med);
  display: flex; align-items: center; justify-content: center;
}
.modal-close:hover { background: #E2EDF4; }

.modal-body { padding: 24px 28px; }
.modal-section { margin-bottom: 24px; }
.modal-section h3 {
  font-size: 0.78rem; color: var(--text-lt); font-weight: 700;
  text-transform: uppercase; letter-spacing: 0.06em; margin: 0 0 10px;
}
.modal-section p { margin: 0 0 8px; color: var(--text-med); line-height: 1.6; }
.formula-box {
  background: var(--surface-2); border-left: 3px solid var(--kalla-primary);
  padding: 12px 16px; border-radius: var(--radius-sm);
  font-family: "SF Mono", Menlo, Consolas, monospace; font-size: 0.86rem;
  color: var(--text-dark);
}

.tabla {
  width: 100%; border-collapse: separate; border-spacing: 0;
  font-size: 0.84rem;
}
.tabla th, .tabla td {
  padding: 8px 10px; border-bottom: 1px solid var(--border); text-align: center;
}
.tabla th {
  background: var(--surface-2); color: var(--text-med);
  font-weight: 600; font-size: 0.76rem; text-transform: uppercase; letter-spacing: 0.04em;
}
.tabla td:first-child, .tabla th:first-child { text-align: left; }
.tabla tbody tr:hover { background: var(--surface-2); }

.year-blocks {
  display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
  gap: 10px;
}
.year-block {
  background: var(--surface-2); border-radius: var(--radius-sm);
  padding: 14px; border-left: 4px solid var(--border);
}
.year-block.cumple  { border-left-color: var(--success); }
.year-block.riesgo  { border-left-color: var(--warning); }
.year-block.critico { border-left-color: var(--danger); }
.year-block.pendiente { border-left-color: var(--neutral); }
.year-block .yr { font-size: 0.76rem; color: var(--text-lt); font-weight: 600; }
.year-block .pct { font-size: 1.4rem; font-weight: 700; color: var(--text-dark); margin: 2px 0; }
.year-block .meta { font-size: 0.78rem; color: var(--text-med); }

.warnings-list {
  background: #FFF7E0; border-left: 3px solid var(--warning);
  border-radius: var(--radius-sm); padding: 10px 14px;
  font-size: 0.82rem; color: #7A5300;
}
.warnings-list ul { margin: 4px 0 0 18px; padding: 0; }

.detail-chart-wrap { position: relative; height: 320px; }

/* ─── Footer ─────────────────────────────────────────────── */
.footer {
  margin-top: 40px; padding: 24px; text-align: center;
  font-size: 0.78rem; color: var(--text-lt);
}

/* ─── Responsive ─────────────────────────────────────────── */
@media (max-width: 720px) {
  .header { padding: 22px 18px; }
  .header h1 { font-size: 1.3rem; }
  .header .meta { text-align: left; }
  .container { padding: 18px; }
  .ind-grid { grid-template-columns: 1fr; }
  .modal { border-radius: var(--radius); }
  .modal-head, .modal-body { padding: 18px; }
  .kpi-grid { grid-template-columns: 1fr 1fr; }
}
.hidden { display: none !important; }
"""


# ============================================================
# JAVASCRIPT
# ============================================================

JS = r"""
// ════════════════════════════════════════════════════════════
// Estado global del dashboard
// ════════════════════════════════════════════════════════════
const STATE = {
  perspFilter: 'all',     // 'all' | 'FINANCIERA' | ...
  yearFilter: 'all',      // 'all' | 2026 | 2027 | ...
  charts: {},             // cache de charts mini
};

const MONTH_ABBR = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"];
const STATUS_LABELS = {
  cumple:    'En meta',
  riesgo:    'En riesgo',
  critico:   'Crítico',
  pendiente: 'Pendiente',
};
const STATUS_COLOR = {
  cumple:    '#23A968',
  riesgo:    '#F0A030',
  critico:   '#E04B4B',
  pendiente: '#94A6B8',
};
const PERSP_KEY_TO_SLUG = {
  'FINANCIERA':'financiera',
  'ASOCIADOS':'asociados',
  'PROCESOS INTERNOS':'procesos',
  'APRENDIZAJE Y CRECIMIENTO':'aprendizaje',
};

// ════════════════════════════════════════════════════════════
// Formateo de valores según tipo de meta
// ════════════════════════════════════════════════════════════
function fmt(v, tipoMeta, unidad) {
  if (v === null || v === undefined || v === '') return '—';
  if (typeof v === 'string') return v;
  if (tipoMeta === 'pct')    return (v*100).toFixed(1) + '%';
  if (tipoMeta === 'bps')    return v.toFixed(0) + ' bps';
  if (tipoMeta === 'ratio')  return ((unidad||'').toLowerCase().includes('x') ? v.toFixed(2)+'x' : v.toFixed(2));
  if (tipoMeta === 'unidades') return Math.round(v).toLocaleString('es');
  return v.toString();
}
function fmtPct(v) {
  if (v === null || v === undefined) return '—';
  return (v*100).toFixed(0) + '%';
}

// ════════════════════════════════════════════════════════════
// KPI cards
// ════════════════════════════════════════════════════════════
function renderKPIs() {
  const s = DATA.summary;
  const el = document.getElementById('kpis');
  const cards = [
    {icon: '📊', label: 'Indicadores', value: s.total, klass: 'total'},
    {icon: '🟢', label: 'En meta',     value: s.cumple || 0, klass: 'cumple'},
    {icon: '🟡', label: 'En riesgo',   value: s.riesgo || 0, klass: 'riesgo'},
    {icon: '🔴', label: 'Crítico',     value: s.critico || 0, klass: 'critico'},
    {icon: '⚪', label: 'Pendiente',   value: s.pendiente || 0, klass: 'pendiente'},
  ];
  el.innerHTML = cards.map(c => `
    <div class="kpi">
      <div class="kpi-icon ${c.klass}">${c.icon}</div>
      <div>
        <div class="kpi-value">${c.value}</div>
        <div class="kpi-label">${c.label}</div>
      </div>
    </div>
  `).join('');
}

// ════════════════════════════════════════════════════════════
// Filtros (chips)
// ════════════════════════════════════════════════════════════
function renderFilters() {
  // Perspectivas
  const persp = document.getElementById('persp-chips');
  const perspButtons = [{key: 'all', name: 'Todas', icon: '◯'}].concat(
    DATA.perspectives.map(p => ({key: p.key, name: p.name, icon: p.icon}))
  );
  persp.innerHTML = perspButtons.map(p => {
    const slug = PERSP_KEY_TO_SLUG[p.key] || '';
    return `<div class="chip persp ${slug ? 'persp-'+slug : ''} ${STATE.perspFilter===p.key?'active':''}" data-persp="${p.key}">
      <span>${p.icon}</span><span>${p.name}</span>
    </div>`;
  }).join('');
  persp.querySelectorAll('.chip').forEach(c => {
    c.addEventListener('click', () => {
      STATE.perspFilter = c.dataset.persp;
      renderFilters();
      renderPerspectives();
    });
  });

  // Años
  const years = document.getElementById('year-chips');
  const yearButtons = [{key: 'all', label: 'Todos los años'}].concat(
    DATA.years.map(y => ({key: y, label: y.toString()}))
  );
  years.innerHTML = yearButtons.map(y => `
    <div class="chip year ${STATE.yearFilter==y.key?'active':''}" data-year="${y.key}">
      ${y.label}
    </div>
  `).join('');
  years.querySelectorAll('.chip').forEach(c => {
    c.addEventListener('click', () => {
      STATE.yearFilter = c.dataset.year === 'all' ? 'all' : parseInt(c.dataset.year);
      renderFilters();
      renderPerspectives();
    });
  });
}

// ════════════════════════════════════════════════════════════
// Render perspectivas e indicadores
// ════════════════════════════════════════════════════════════
function renderPerspectives() {
  const root = document.getElementById('perspectives');
  // limpiar mini-charts antes de regenerar
  Object.values(STATE.charts).forEach(ch => { try { ch.destroy(); } catch(e){} });
  STATE.charts = {};

  const visiblePersps = DATA.perspectives.filter(p =>
    STATE.perspFilter === 'all' || STATE.perspFilter === p.key
  );

  if (visiblePersps.length === 0) {
    root.innerHTML = '<div style="text-align:center;padding:40px;color:var(--text-lt)">No hay perspectivas que coincidan con el filtro.</div>';
    return;
  }

  root.innerHTML = visiblePersps.map(p => {
    const slug = PERSP_KEY_TO_SLUG[p.key] || '';
    const cards = p.indicators.map(ind => renderCard(ind, p)).join('');
    return `
      <section class="perspective">
        <div class="persp-header" style="background:linear-gradient(135deg, ${p.accent}, ${shade(p.accent, -15)})">
          <span class="icon">${p.icon}</span>
          <span class="title">Perspectiva ${p.name}</span>
          <span class="count">${p.indicators.length} indicadores</span>
        </div>
        <div class="persp-objetivo" style="color: ${p.accent}">
          <strong>Objetivo general:</strong> <span style="color:var(--text-med)">${escapeHTML(p.objetivo_general)}</span>
        </div>
        <div class="ind-grid">${cards}</div>
      </section>
    `;
  }).join('');

  // Generar mini-charts después de inyectar el DOM
  visiblePersps.forEach(p => {
    p.indicators.forEach(ind => buildMiniChart(ind));
  });

  // Click en card → abrir modal
  root.querySelectorAll('.ind-card').forEach(card => {
    card.addEventListener('click', () => openModal(card.dataset.code));
  });
}

function renderCard(ind, persp) {
  // Métricas según filtro de año
  let displayYearObj = null;
  if (STATE.yearFilter !== 'all') {
    displayYearObj = ind.yearly.find(y => y.year === STATE.yearFilter);
  } else {
    // último año con datos
    for (let i = ind.yearly.length-1; i>=0; i--) {
      if (ind.yearly[i].cumplimiento !== null) { displayYearObj = ind.yearly[i]; break; }
    }
    if (!displayYearObj) displayYearObj = ind.yearly[0];
  }

  const status = displayYearObj.status;
  const meta = fmt(displayYearObj.meta_raw, ind.tipo_meta, ind.unidad);
  const result = fmt(displayYearObj.avg_result, ind.tipo_meta, ind.unidad);
  const cumpl = fmtPct(displayYearObj.cumplimiento);
  const yr = displayYearObj.year;

  const trendIcon = ind.trend === 'up' ? '↑' : ind.trend === 'down' ? '↓' : ind.trend === 'flat' ? '→' : '·';
  const trendClass = ind.trend ? 'trend-' + ind.trend : '';
  const accumulated = ind.accumulated !== null ? fmtPct(ind.accumulated) : '—';
  const alert = ind.alert || {level: 'pending', text: 'Sin datos'};

  return `
    <div class="ind-card" data-code="${ind.code}">
      <div class="accent-bar" style="background:${persp.accent}"></div>
      <div class="ind-head">
        <div style="min-width:0;flex:1">
          <div class="ind-code">${ind.code}</div>
          <h3 class="ind-name">${escapeHTML(ind.name)}</h3>
          <div class="ind-ini">${ind.ini_icon} ${escapeHTML(ind.ini)}</div>
        </div>
        <span class="status-pill ${status}">${STATUS_LABELS[status]}</span>
      </div>

      <div class="ind-metrics">
        <div class="metric">
          <div class="label">Meta ${yr}</div>
          <div class="value">${meta}</div>
        </div>
        <div class="metric">
          <div class="label">Resultado</div>
          <div class="value">${result}</div>
        </div>
        <div class="metric">
          <div class="label">Cumpl. ${yr}</div>
          <div class="value ${status}">${cumpl}</div>
        </div>
      </div>

      <div class="chart-wrap"><canvas id="chart-${ind.code}"></canvas></div>

      <div class="ind-foot">
        <div class="foot-cell">
          <span class="l">Tendencia</span>
          <span class="v ${trendClass}">${trendIcon} ${ind.tend}</span>
        </div>
        <div class="foot-cell">
          <span class="l">Avance global</span>
          <span class="v">${accumulated}</span>
        </div>
        <div class="foot-cell">
          <span class="l">Responsable</span>
          <span class="v">${escapeHTML(ind.responsable || 'Por definir')}</span>
        </div>
      </div>

      <div class="alert ${alert.level}">⚠ ${escapeHTML(alert.text)}</div>
    </div>
  `;
}

// ════════════════════════════════════════════════════════════
// Mini-chart (sparkline) por indicador
// ════════════════════════════════════════════════════════════
function buildMiniChart(ind) {
  const ctx = document.getElementById('chart-' + ind.code);
  if (!ctx) return;
  const labels = DATA.months_labels;
  const data = ind.monthly_results.map(r => r === null ? null : r);
  // Línea de meta: array de 60 valores con meta del año correspondiente
  const metaLine = [];
  for (let i = 0; i < 60; i++) {
    const yi = Math.floor(i / 12);
    const meta = ind.metas[yi];
    metaLine.push(typeof meta === 'number' ? meta : null);
  }
  const persp = DATA.perspectives.find(p => p.indicators.includes(ind));
  const color = persp ? persp.accent : '#0876B2';

  const chart = new Chart(ctx, {
    type: 'line',
    data: {
      labels: labels,
      datasets: [
        {
          label: 'Resultado',
          data: data,
          borderColor: color,
          backgroundColor: hexA(color, 0.12),
          tension: 0.3, fill: true,
          pointRadius: 0, pointHoverRadius: 4,
          spanGaps: true, borderWidth: 2,
        },
        {
          label: 'Meta',
          data: metaLine,
          borderColor: '#94A6B8',
          borderDash: [4, 4],
          borderWidth: 1.4,
          pointRadius: 0, fill: false,
        }
      ]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: {
        legend: { display: false },
        tooltip: {
          mode: 'index', intersect: false,
          callbacks: {
            label: (ctx) => {
              if (ctx.raw === null || ctx.raw === undefined) return ctx.dataset.label + ': —';
              return ctx.dataset.label + ': ' + fmt(ctx.raw, ind.tipo_meta, ind.unidad);
            }
          }
        }
      },
      scales: {
        x: { display: false },
        y: { display: false, beginAtZero: false }
      }
    }
  });
  STATE.charts[ind.code] = chart;
}

// ════════════════════════════════════════════════════════════
// Modal con detalle del indicador
// ════════════════════════════════════════════════════════════
function openModal(code) {
  const ind = findIndicator(code);
  if (!ind) return;
  const persp = DATA.perspectives.find(p => p.indicators.includes(ind));
  const overlay = document.getElementById('modal-overlay');
  overlay.classList.add('active');

  // Año blocks
  const yearBlocks = ind.yearly.map(y => `
    <div class="year-block ${y.status}">
      <div class="yr">${y.year}</div>
      <div class="pct">${y.cumplimiento !== null ? (y.cumplimiento*100).toFixed(0)+'%' : '—'}</div>
      <div class="meta">Meta: ${fmt(y.meta_raw, ind.tipo_meta, ind.unidad)}</div>
      <div class="meta">Resultado: ${fmt(y.avg_result, ind.tipo_meta, ind.unidad)}</div>
    </div>
  `).join('');

  // Tabla mensual
  let tabla = `<table class="tabla">
    <thead><tr><th>Año</th>${MONTH_ABBR.map(m=>`<th>${m}</th>`).join('')}</tr></thead>
    <tbody>`;
  for (let yi = 0; yi < 5; yi++) {
    const yr = 2026 + yi;
    let row = `<tr><td><strong>${yr}</strong></td>`;
    for (let m = 0; m < 12; m++) {
      const v = ind.monthly_results[yi*12 + m];
      row += `<td>${v === null ? '<span style="color:var(--text-lt)">—</span>' : fmt(v, ind.tipo_meta, ind.unidad)}</td>`;
    }
    row += '</tr>';
    tabla += row;
  }
  tabla += '</tbody></table>';

  const warningsHTML = ind.warnings && ind.warnings.length ? `
    <div class="warnings-list">
      <strong>⚠ Avisos de validación:</strong>
      <ul>${ind.warnings.map(w => `<li>${escapeHTML(w)}</li>`).join('')}</ul>
    </div>` : '';

  document.getElementById('modal-content').innerHTML = `
    <div class="modal-head">
      <div class="titles">
        <div class="code">${ind.code}</div>
        <div class="name">${escapeHTML(ind.name)}</div>
        <div class="persp-tag">${persp.icon} ${persp.name} · ${ind.ini_icon} ${escapeHTML(ind.ini)}</div>
      </div>
      <button class="modal-close" id="modal-close">×</button>
    </div>
    <div class="modal-body">
      <div class="modal-section">
        <h3>📐 Fórmula</h3>
        <div class="formula-box">${escapeHTML(ind.formula)}</div>
      </div>
      <div class="modal-section">
        <h3>📝 Descripción</h3>
        <p>${escapeHTML(ind.desc)}</p>
      </div>
      <div class="modal-section" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px">
        <div><h3>Unidad</h3><p style="color:var(--text-dark);font-weight:600">${escapeHTML(ind.unidad)}</p></div>
        <div><h3>Tendencia objetivo</h3><p style="color:var(--text-dark);font-weight:600">${ind.tend}</p></div>
        <div><h3>Frecuencia</h3><p style="color:var(--text-dark);font-weight:600">${ind.frec}</p></div>
        <div><h3>Responsable</h3><p style="color:var(--text-dark);font-weight:600">${escapeHTML(ind.responsable)}</p></div>
      </div>
      ${warningsHTML}
      <div class="modal-section">
        <h3>📅 Cumplimiento por año</h3>
        <div class="year-blocks">${yearBlocks}</div>
      </div>
      <div class="modal-section">
        <h3>📈 Resultados mensuales (Ene 2026 → Dic 2030)</h3>
        <div class="detail-chart-wrap"><canvas id="detail-chart"></canvas></div>
      </div>
      <div class="modal-section">
        <h3>📋 Sábana mensual</h3>
        <div style="overflow-x:auto">${tabla}</div>
      </div>
    </div>
  `;
  document.getElementById('modal-close').addEventListener('click', closeModal);
  overlay.addEventListener('click', e => { if (e.target === overlay) closeModal(); });
  // chart detallado
  setTimeout(() => buildDetailChart(ind, persp.accent), 50);
}

function buildDetailChart(ind, color) {
  const ctx = document.getElementById('detail-chart');
  if (!ctx) return;
  if (STATE.charts.detail) try { STATE.charts.detail.destroy(); } catch(e){}
  const metaLine = [];
  for (let i = 0; i < 60; i++) {
    const yi = Math.floor(i / 12);
    const meta = ind.metas[yi];
    metaLine.push(typeof meta === 'number' ? meta : null);
  }
  STATE.charts.detail = new Chart(ctx, {
    type: 'line',
    data: {
      labels: DATA.months_labels,
      datasets: [
        {
          label: 'Resultado mensual',
          data: ind.monthly_results,
          borderColor: color,
          backgroundColor: hexA(color, 0.14),
          tension: 0.25, fill: true,
          pointRadius: 2.5, pointHoverRadius: 5,
          spanGaps: true, borderWidth: 2.2,
        },
        {
          label: 'Meta anual',
          data: metaLine,
          borderColor: '#94A6B8',
          borderDash: [6, 4],
          borderWidth: 1.6,
          pointRadius: 0, fill: false,
        }
      ]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      interaction: { mode: 'index', intersect: false },
      plugins: {
        legend: { position: 'bottom', labels: { font: { size: 12 } } },
        tooltip: {
          callbacks: {
            label: (ctx) => {
              if (ctx.raw === null || ctx.raw === undefined) return ctx.dataset.label + ': sin dato';
              return ctx.dataset.label + ': ' + fmt(ctx.raw, ind.tipo_meta, ind.unidad);
            }
          }
        }
      },
      scales: {
        x: { ticks: { autoSkip: true, maxRotation: 0, font: { size: 10 } }, grid: { color: '#EEF3F7' } },
        y: { ticks: { callback: (v) => fmt(v, ind.tipo_meta, ind.unidad), font: { size: 11 } }, grid: { color: '#EEF3F7' } }
      }
    }
  });
}

function closeModal() {
  document.getElementById('modal-overlay').classList.remove('active');
  if (STATE.charts.detail) { try { STATE.charts.detail.destroy(); } catch(e){} delete STATE.charts.detail; }
}

// ════════════════════════════════════════════════════════════
// Helpers
// ════════════════════════════════════════════════════════════
function findIndicator(code) {
  for (const p of DATA.perspectives) {
    for (const ind of p.indicators) {
      if (ind.code === code) return ind;
    }
  }
  return null;
}
function escapeHTML(s) {
  if (s === null || s === undefined) return '';
  return String(s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'})[c]);
}
function hexA(hex, alpha) {
  const h = hex.replace('#','');
  const r = parseInt(h.substring(0,2),16), g = parseInt(h.substring(2,4),16), b = parseInt(h.substring(4,6),16);
  return `rgba(${r},${g},${b},${alpha})`;
}
function shade(hex, percent) {
  const h = hex.replace('#','');
  let r = parseInt(h.substring(0,2),16), g = parseInt(h.substring(2,4),16), b = parseInt(h.substring(4,6),16);
  const f = percent / 100;
  r = Math.max(0, Math.min(255, Math.round(r + 255*f)));
  g = Math.max(0, Math.min(255, Math.round(g + 255*f)));
  b = Math.max(0, Math.min(255, Math.round(b + 255*f)));
  return '#' + [r,g,b].map(v => v.toString(16).padStart(2,'0')).join('');
}

// ════════════════════════════════════════════════════════════
// Inicialización
// ════════════════════════════════════════════════════════════
document.addEventListener('DOMContentLoaded', () => {
  renderKPIs();
  renderFilters();
  renderPerspectives();
  document.addEventListener('keydown', e => { if (e.key === 'Escape') closeModal(); });
});
"""


# ============================================================
# HTML TEMPLATE
# ============================================================

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Plan Estratégico 2026–2030 · Dashboard</title>
  <meta name="description" content="Dashboard de seguimiento de indicadores del Plan Estratégico 2026–2030.">
  <script>__CHARTJS__</script>
  <style>
__CSS__
  </style>
</head>
<body>
  <header class="header">
    <div class="header-inner">
      <div>
        <h1><span class="brand-dot"></span>Plan Estratégico 2026 – 2030</h1>
        <div class="sub">Dashboard de seguimiento · 25 indicadores · 4 perspectivas estratégicas</div>
      </div>
      <div class="meta">
        <div>Última actualización</div>
        <div><strong>__GENERATED_AT__</strong></div>
      </div>
    </div>
  </header>

  <div class="container">
    <div id="kpis" class="kpi-grid"></div>

    <div class="filters">
      <div class="filter-group">
        <span class="filter-label">Perspectiva:</span>
        <div id="persp-chips" style="display:flex;gap:8px;flex-wrap:wrap"></div>
      </div>
      <div class="filter-group">
        <span class="filter-label">Año:</span>
        <div id="year-chips" style="display:flex;gap:8px;flex-wrap:wrap"></div>
      </div>
    </div>

    <div id="perspectives"></div>

    <div class="footer">
      Generado automáticamente desde Tablero_Indicadores_PE.xlsx · Paleta inspirada en KALLA<br>
      Para actualizar: ejecute <code>python3 "Plan Estrategico Dashboard.py"</code> después de modificar el Excel.
    </div>
  </div>

  <div class="modal-overlay" id="modal-overlay">
    <div class="modal" id="modal-content" onclick="event.stopPropagation()"></div>
  </div>

  <script>
    const DATA = __DATA_JSON__;
__JS__
  </script>
</body>
</html>
"""


# ============================================================
# RENDER
# ============================================================

def _load_chartjs():
    """Carga Chart.js desde el archivo local si existe; si no, deja un fallback
    que carga desde CDN (útil cuando el dashboard se publica en GitHub Pages)."""
    import os
    here = os.path.dirname(os.path.abspath(__file__))
    chart_path = os.path.join(here, 'chart_umd.js')
    if os.path.exists(chart_path):
        with open(chart_path, 'r', encoding='utf-8') as f:
            return f.read()
    # Fallback: inyectar un loader que carga Chart.js desde CDN
    return (
        "/* Chart.js no embebido — cargando desde CDN */\n"
        "(function(){var s=document.createElement('script');"
        "s.src='https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js';"
        "s.onload=function(){console.log('Chart.js CDN cargado.');};"
        "document.head.appendChild(s);})();"
    )


def render_dashboard_html(payload):
    """Renderiza el dashboard como HTML autocontenido (incluye Chart.js inline)."""
    json_str = json.dumps(payload, ensure_ascii=False, indent=2)
    html = (HTML_TEMPLATE
            .replace('__CHARTJS__', _load_chartjs())
            .replace('__CSS__', CSS)
            .replace('__JS__', JS)
            .replace('__DATA_JSON__', json_str)
            .replace('__GENERATED_AT__', payload['generated_at']))
    return html


# ============================================================================
# 3-8. EXTRACCIÓN, CÁLCULO, VALIDACIÓN, PAYLOAD, RENDER Y MAIN
# ============================================================================


# openpyxl es la única dependencia externa
# Imports locales
# (data_model y dashboard_templates están incluidos arriba en este mismo archivo)


# ============================================================
# CONFIGURACIÓN
# ============================================================

YEARS = [2026, 2027, 2028, 2029, 2030]
MONTHS = []
for y in YEARS:
    for m in range(1, 13):
        MONTHS.append((y, m))

MES_NAMES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
MES_ABBR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def freq_months(frec):
    """Devuelve el conjunto de meses (1-12) en que se mide el indicador."""
    f = (frec or "").lower()
    if "mensual" in f:
        return set(range(1, 13))
    if "trimestral" in f:
        return {3, 6, 9, 12}
    if "semestral" in f:
        return {6, 12}
    if "anual" in f:
        return {12}
    return set(range(1, 13))


# ============================================================
# EXTRACCIÓN DE DATOS DEL EXCEL
# ============================================================

def _find_data_first_row(ws):
    """Localiza la primera fila de la tabla mensual (donde col B = 1)."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value == 1:
            return r
    return None


def _read_metas(ws, data_first):
    """Lee las 5 metas anuales desde la tabla de metas (sobre data_first)."""
    metas = []
    for r in range(1, data_first):
        if ws.cell(r, 2).value == 2026:
            for i in range(5):
                metas.append(ws.cell(r + i, 3).value)
            break
    return metas


def _read_responsables(wb):
    """Extrae el catálogo de responsables de la hoja Parámetros."""
    responsables = {}
    if 'Parámetros' not in wb.sheetnames:
        return responsables
    pws = wb['Parámetros']
    for r in range(1, pws.max_row + 1):
        if pws.cell(r, 2).value == 'Código':
            for rr in range(r + 1, r + 1 + len(IND) + 5):
                code = pws.cell(rr, 2).value
                resp = pws.cell(rr, 6).value
                if code and isinstance(code, str) and code.startswith('IND-'):
                    responsables[code] = resp
            break
    return responsables


def extract_indicator(wb, d):
    """Extrae datos crudos de la hoja de un indicador."""
    sheet_name = d['code']
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠️  Hoja {sheet_name} no encontrada — se omite")
        return None
    ws = wb[sheet_name]

    data_first = _find_data_first_row(ws)
    if data_first is None:
        print(f"  ⚠️  No se localizó la tabla mensual en {sheet_name}")
        return None

    metas_raw = _read_metas(ws, data_first)
    if len(metas_raw) != 5:
        metas_raw = list(d['metas'])  # fallback al catálogo

    # Leer variables mensuales (columnas E, F, G, ... según número de vars)
    nv = len(d['vars'])
    var_col_start = 5  # columna E
    inputs = []
    for i, (y, m) in enumerate(MONTHS):
        row = data_first + i
        rec = {'year': y, 'month': m}
        for j, (var_code, _) in enumerate(d['vars']):
            v = ws.cell(row, var_col_start + j).value
            rec[var_code] = v if isinstance(v, (int, float)) else None
        inputs.append(rec)

    return {
        'code': d['code'],
        'name': d['indicador'],
        'persp': d['p'],
        'ini': d['ini'],
        'formula': d['formula'],
        'desc': d['desc'],
        'unidad': d['unidad'],
        'tipo': d['tipo'],
        'tipo_meta': d['tipo_meta'],
        'tend': d['tend'],
        'frec': d['frec'],
        'metas_raw': metas_raw,
        'inputs': inputs,
        'vars': d['vars'],
        'allowed_months': sorted(freq_months(d['frec'])),
    }


# ============================================================
# CÁLCULOS
# ============================================================

def compute_monthly_results(ind):
    """Calcula el resultado de cada uno de los 60 meses según el tipo de fórmula."""
    tipo = ind['tipo']
    inputs = ind['inputs']
    allowed = set(ind['allowed_months'])
    results = []

    for i, rec in enumerate(inputs):
        if rec['month'] not in allowed:
            results.append(None)
            continue

        var_keys = [k for k, _ in ind['vars']]
        result = None

        if tipo == 'ratio':
            # Último var = denominador; los anteriores se suman como numerador
            den = rec[var_keys[-1]]
            num_vals = [rec[k] for k in var_keys[:-1]]
            if den is not None and den != 0 and all(v is not None for v in num_vals):
                result = sum(num_vals) / den

        elif tipo == 'growth':
            cur = rec[var_keys[0]]
            if i == 0 or cur is None:
                result = None
            else:
                prev = inputs[i - 1][var_keys[0]]
                if prev is not None and prev != 0:
                    result = (cur - prev) / prev

        elif tipo == 'diff':
            a, b = rec[var_keys[0]], rec[var_keys[1]]
            if a is not None and b is not None:
                diff = a - b
                # Para indicadores en bps, multiplicar por 10000
                result = diff * 10000 if ind['tipo_meta'] == 'bps' else diff

        elif tipo in ('avg', 'count'):
            result = rec[var_keys[0]]

        results.append(result)

    return results


def get_meta_numeric(ind, year):
    """Devuelve la meta numérica del año, o None si es texto."""
    m = ind['metas_raw'][YEARS.index(year)]
    return m if isinstance(m, (int, float)) else None


def compute_compliance(result, meta, tend):
    """% de cumplimiento ajustado por tendencia."""
    if result is None or meta is None:
        return None
    if tend == 'Reducir':
        if result == 0:
            return None
        return meta / result
    if tend == 'Mantener':
        return 1 - abs(result - meta) / max(abs(meta), 1e-9)
    # Crecer
    if meta == 0:
        return None
    return result / meta


def status_from_compliance(c):
    """Mapea cumplimiento a estado."""
    if c is None:
        return 'pendiente'
    if c >= 0.95:
        return 'cumple'
    if c >= 0.80:
        return 'riesgo'
    return 'critico'


def compute_yearly_metrics(ind, monthly_results):
    """Para cada año (2026-2030): meta, resultado promedio, % cumplimiento, estado."""
    out = []
    for yi, year in enumerate(YEARS):
        seg = monthly_results[yi * 12: (yi + 1) * 12]
        valid = [r for r in seg if r is not None]
        meta = get_meta_numeric(ind, year)

        avg_result = sum(valid) / len(valid) if valid else None

        # Promedio de cumplimientos mensuales del año
        cmpls = []
        for r in valid:
            c = compute_compliance(r, meta, ind['tend'])
            if c is not None:
                cmpls.append(c)
        cumpl = sum(cmpls) / len(cmpls) if cmpls else None

        out.append({
            'year': year,
            'meta': meta,
            'meta_raw': ind['metas_raw'][yi],
            'avg_result': avg_result,
            'cumplimiento': cumpl,
            'status': status_from_compliance(cumpl),
            'n_months': len(valid),
        })
    return out


def compute_trend(monthly_results, ind):
    """Tendencia propia del indicador: 'up' / 'down' / 'flat' / None.

    Se calcula comparando el primer y último resultado disponibles en
    los últimos 12 meses con datos.
    """
    valid = [r for r in monthly_results if r is not None]
    if len(valid) < 2:
        return None
    window = valid[-12:] if len(valid) >= 12 else valid
    first, last = window[0], window[-1]
    if abs(last - first) / max(abs(first), 1e-9) < 0.02:
        return 'flat'
    return 'up' if last > first else 'down'


def compute_accumulated(monthly_results, ind):
    """Avance acumulado: cumplimiento del último resultado disponible vs meta 2030."""
    meta_2030 = get_meta_numeric(ind, 2030)
    if meta_2030 is None:
        return None
    latest = next((r for r in reversed(monthly_results) if r is not None), None)
    if latest is None:
        return None
    return compute_compliance(latest, meta_2030, ind['tend'])


def compute_alert(yearly_metrics):
    """Genera la alerta de desviación basada en el último año con datos."""
    for ym in reversed(yearly_metrics):
        if ym['cumplimiento'] is not None:
            c = ym['cumplimiento']
            y = ym['year']
            if c >= 0.95:
                return {'level': 'ok', 'text': f'En meta en {y} ({c*100:.0f}%)'}
            if c >= 0.80:
                return {'level': 'warn', 'text': f'Atención en {y}: {c*100:.0f}% de la meta'}
            return {'level': 'critical', 'text': f'Crítico en {y}: {c*100:.0f}% — revisar'}
    return {'level': 'pending', 'text': 'Sin datos cargados aún'}


# ============================================================
# VALIDACIÓN
# ============================================================

def validate_indicator(ind):
    """Valida los datos de un indicador y devuelve una lista de avisos."""
    warnings = []
    # ¿Las metas tienen el tipo esperado?
    for i, m in enumerate(ind['metas_raw']):
        if m is None or (isinstance(m, str) and 'definir' in m.lower()):
            warnings.append(f"Meta {YEARS[i]} sin definir")
    # ¿Hay meses con datos parciales (algunas variables sí, otras no)?
    var_keys = [k for k, _ in ind['vars']]
    for rec in ind['inputs']:
        if rec['month'] not in set(ind['allowed_months']):
            continue
        present = [rec[k] for k in var_keys]
        n_filled = sum(1 for v in present if v is not None)
        if 0 < n_filled < len(var_keys):
            warnings.append(f"Datos parciales en {MES_ABBR[rec['month']-1]} {rec['year']}")
    return warnings


# ============================================================
# PROCESAMIENTO COMPLETO
# ============================================================

def process_workbook(xlsx_path):
    """Procesa el workbook completo y devuelve la lista de indicadores enriquecidos."""
    wb = load_workbook(xlsx_path, data_only=False)
    responsables = _read_responsables(wb)

    indicators = []
    for d in IND:
        ind = extract_indicator(wb, d)
        if ind is None:
            continue
        ind['responsable'] = responsables.get(ind['code']) or 'Por definir'
        ind['monthly_results'] = compute_monthly_results(ind)
        ind['yearly'] = compute_yearly_metrics(ind, ind['monthly_results'])
        ind['trend'] = compute_trend(ind['monthly_results'], ind)
        ind['accumulated'] = compute_accumulated(ind['monthly_results'], ind)
        ind['alert'] = compute_alert(ind['yearly'])
        ind['warnings'] = validate_indicator(ind)
        indicators.append(ind)
    return indicators


# ============================================================
# CONSTRUCCIÓN DE PAYLOAD PARA EL DASHBOARD
# ============================================================

def _to_jsonable(v):
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    if isinstance(v, (int, str, bool)):
        return v
    return str(v)


def build_payload(indicators):
    """Construye el JSON consolidado que consume el dashboard."""
    perspectives = {}
    for p in PERSP_ORDER:
        perspectives[p] = {
            'key': p,
            'name': PERSP_INFO[p]['short'],
            'icon': PERSP_INFO[p]['icon'],
            'accent': '#' + PERSP_INFO[p]['accent'],
            'objetivo_general': OBJ_GEN[p],
            'indicators': [],
        }

    summary = {'total': 0, 'cumple': 0, 'riesgo': 0, 'critico': 0, 'pendiente': 0}

    for ind in indicators:
        # Estado consolidado del indicador (último año con datos)
        latest_status = 'pendiente'
        latest_cumpl = None
        latest_year = None
        for ym in reversed(ind['yearly']):
            if ym['cumplimiento'] is not None:
                latest_status = ym['status']
                latest_cumpl = ym['cumplimiento']
                latest_year = ym['year']
                break

        payload_ind = {
            'code': ind['code'],
            'name': ind['name'],
            'ini': ind['ini'],
            'ini_icon': INI_ICON.get(ind['ini'], '🎯'),
            'ini_desc': INI_DESC.get(ind['ini'], ''),
            'formula': ind['formula'],
            'desc': ind['desc'],
            'unidad': ind['unidad'],
            'tipo': ind['tipo'],
            'tipo_meta': ind['tipo_meta'],
            'tend': ind['tend'],
            'frec': ind['frec'],
            'responsable': ind['responsable'],
            'vars': [{'code': v[0], 'label': v[1]} for v in ind['vars']],
            'metas': [_to_jsonable(m) for m in ind['metas_raw']],
            'monthly_results': [_to_jsonable(r) for r in ind['monthly_results']],
            'yearly': [{
                'year': y['year'],
                'meta': _to_jsonable(y['meta']),
                'meta_raw': _to_jsonable(y['meta_raw']),
                'avg_result': _to_jsonable(y['avg_result']),
                'cumplimiento': _to_jsonable(y['cumplimiento']),
                'status': y['status'],
                'n_months': y['n_months'],
            } for y in ind['yearly']],
            'trend': ind['trend'],
            'accumulated': _to_jsonable(ind['accumulated']),
            'alert': ind['alert'],
            'warnings': ind['warnings'],
            'allowed_months': ind['allowed_months'],
            'latest_status': latest_status,
            'latest_cumpl': _to_jsonable(latest_cumpl),
            'latest_year': latest_year,
        }
        perspectives[ind['persp']]['indicators'].append(payload_ind)
        summary['total'] += 1
        summary[latest_status] = summary.get(latest_status, 0) + 1

    return {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'months_labels': [f"{MES_ABBR[m-1]} {y}" for (y, m) in MONTHS],
        'years': YEARS,
        'perspectives': list(perspectives.values()),
        'summary': summary,
    }


# ============================================================
# AUTOPUBLICACIÓN A GITHUB
# ============================================================

def _github_api(method, path, token, data=None, timeout=20):
    """Llamada genérica a la GitHub REST API. Devuelve (status, json_dict)."""
    import urllib.request
    import urllib.error
    url = f'https://api.github.com{path}'
    body = None
    if data is not None:
        body = json.dumps(data).encode('utf-8')
    req = urllib.request.Request(
        url,
        data=body,
        method=method,
        headers={
            'Accept': 'application/vnd.github+json',
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
            'X-GitHub-Api-Version': '2022-11-28',
            'User-Agent': 'Plan-Estrategico-Dashboard',
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            raw = r.read().decode('utf-8') or '{}'
            return r.status, json.loads(raw)
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode('utf-8')
            err_json = json.loads(err_body) if err_body else {}
        except Exception:
            err_json = {'message': str(e)}
        return e.code, err_json


def setup_github(token, repo_name, git_root, out_path):
    """Configuración inicial completa: crea el repo en GitHub, configura remote,
    hace push inicial y activa GitHub Pages. Todo vía API + git CLI."""
    import subprocess

    print(f"\n🔧 Configurando GitHub...")

    # 1. Validar token y obtener usuario
    status, user_info = _github_api('GET', '/user', token)
    if status != 200:
        print(f"   ❌ Token inválido o sin permisos ({status}).")
        print(f"      Verifique que el token tiene el scope 'repo'.")
        print(f"      Genere uno nuevo en: https://github.com/settings/tokens/new")
        return None
    username = user_info['login']
    print(f"   ✓ Autenticado como: {username}")

    # 2. Verificar si el repo existe; crear si no
    status, _ = _github_api('GET', f'/repos/{username}/{repo_name}', token)
    if status == 404:
        print(f"   ⏳ Creando repo {username}/{repo_name}...")
        status, resp = _github_api('POST', '/user/repos', token, data={
            'name': repo_name,
            'private': False,
            'auto_init': False,
            'description': 'Dashboard del Plan Estratégico 2026-2030',
            'has_issues': False,
            'has_wiki': False,
        })
        if status not in (200, 201):
            print(f"   ❌ No se pudo crear el repo: {resp.get('message', 'error')}")
            for err in resp.get('errors', []):
                print(f"      - {err.get('message', err)}")
            return None
        print(f"   ✓ Repo creado: https://github.com/{username}/{repo_name}")
    elif status == 200:
        print(f"   ✓ Repo existente: {username}/{repo_name}")
    else:
        print(f"   ❌ Error consultando el repo ({status})")
        return None

    # 3. Configurar remote 'origin' con token embebido (queda sólo en .git/config local)
    remote_url = f'https://x-access-token:{token}@github.com/{username}/{repo_name}.git'
    safe_url = f'https://github.com/{username}/{repo_name}.git'
    r = subprocess.run(['git', 'remote'], capture_output=True, text=True, cwd=str(git_root))
    remotes = r.stdout.split()
    if 'origin' in remotes:
        subprocess.run(['git', 'remote', 'set-url', 'origin', remote_url], cwd=str(git_root))
        print(f"   ✓ Remote 'origin' actualizado")
    else:
        subprocess.run(['git', 'remote', 'add', 'origin', remote_url], cwd=str(git_root))
        print(f"   ✓ Remote 'origin' agregado: {safe_url}")

    # 4. Asegurar que la rama sea 'main'
    subprocess.run(['git', 'branch', '-M', 'main'], capture_output=True, cwd=str(git_root))

    # 5. Si no hay commits aún, hacer commit inicial con el dashboard
    r = subprocess.run(['git', 'log', '-1', '--oneline'], capture_output=True, text=True, cwd=str(git_root))
    if r.returncode != 0:
        rel = str(out_path.resolve().relative_to(git_root))
        subprocess.run(['git', 'add', rel], cwd=str(git_root))
        subprocess.run(['git', 'commit', '-m', 'Inicial: dashboard del Plan Estratégico'],
                       capture_output=True, cwd=str(git_root))
        print(f"   ✓ Commit inicial creado")

    # 6. Push inicial
    print(f"   ⏳ Push inicial a GitHub...")
    r = subprocess.run(['git', 'push', '-u', 'origin', 'main'],
                       capture_output=True, text=True, cwd=str(git_root))
    if r.returncode != 0:
        err = r.stderr.strip() or r.stdout.strip()
        # Ocultar el token si aparece en el error
        err = err.replace(token, '***')
        print(f"   ❌ Push falló:")
        for line in err.split('\n')[:6]:
            print(f"      {line}")
        return None
    print(f"   ✓ Push exitoso")

    # 7. Activar GitHub Pages (branch main, folder /docs)
    print(f"   ⏳ Activando GitHub Pages...")
    status, resp = _github_api('POST', f'/repos/{username}/{repo_name}/pages', token, data={
        'source': {'branch': 'main', 'path': '/docs'},
    })
    if status in (200, 201):
        print(f"   ✓ GitHub Pages activado (branch main, folder /docs)")
    elif status == 409:
        print(f"   ℹ️  GitHub Pages ya estaba activado")
    else:
        print(f"   ⚠️  No se pudo activar Pages automáticamente ({status}: {resp.get('message','')})")
        print(f"      Actívelo manualmente en:")
        print(f"      https://github.com/{username}/{repo_name}/settings/pages")

    pages_url = f'https://{username}.github.io/{repo_name}/'
    print(f"\n🎉 ¡Configuración completa!")
    print(f"   📦 Repo:      https://github.com/{username}/{repo_name}")
    print(f"   🌐 Dashboard: {pages_url}")
    print(f"   ⏱️  La URL estará disponible en 1-2 minutos.")
    print(f"\n   Desde ahora, cada `python \"Plan Estrategico Dashboard.py\"`")
    print(f"   actualizará el dashboard publicado automáticamente.")
    return pages_url


def _find_git_root(start_path):
    """Busca la raíz del repo Git subiendo desde start_path."""
    p = Path(start_path).resolve()
    for _ in range(8):
        if (p / '.git').exists():
            return p
        if p.parent == p:
            break
        p = p.parent
    return None


def _try_gh_setup(git_root, out_path):
    """Intenta crear el repo + push + Pages usando GitHub CLI (gh).

    Retorna True si tuvo éxito. False si gh no está instalado o no está
    autenticado (el caller decidirá qué hacer).
    """
    import subprocess

    # 1. ¿gh está instalado?
    try:
        r = subprocess.run(['gh', '--version'], capture_output=True, text=True, timeout=10)
        if r.returncode != 0:
            raise FileNotFoundError
    except (FileNotFoundError, subprocess.TimeoutExpired):
        print(f"\n💡 GitHub CLI no está instalado en este equipo.")
        print(f"   Para publicación 100% automática (recomendado):")
        print(f"      1. Instale gh:  https://cli.github.com  (instalador para Windows)")
        print(f"      2. En la terminal:  gh auth login")
        print(f"      3. Ejecute de nuevo: python \"Plan Estrategico Dashboard.py\"")
        print(f"\n   Alternativa con Personal Access Token:")
        print(f"      python \"Plan Estrategico Dashboard.py\" --setup-github")
        return False

    # 2. ¿gh está autenticado?
    r = subprocess.run(['gh', 'auth', 'status'], capture_output=True, text=True)
    if r.returncode != 0:
        print(f"\n💡 GitHub CLI instalado pero no autenticado.")
        print(f"   Ejecute una vez:    gh auth login")
        print(f"   Después:             python \"Plan Estrategico Dashboard.py\"")
        return False

    # 3. Obtener el nombre de usuario
    r = subprocess.run(['gh', 'api', 'user', '-q', '.login'],
                       capture_output=True, text=True)
    username = r.stdout.strip()
    if not username:
        print(f"   ⚠️  No se pudo obtener el usuario de GitHub")
        return False

    # 4. Nombre del repo (= nombre de carpeta sanitizado)
    repo_name = git_root.name.lower()
    repo_name = ''.join(c if c.isalnum() or c in '-_' else '-' for c in repo_name)
    repo_name = repo_name.strip('-') or 'plan-estrategico'

    print(f"\n🔧 Configuración automática vía GitHub CLI")
    print(f"   Usuario: {username}")
    print(f"   Repo:    {repo_name}")

    # 5. Asegurar branch 'main'
    subprocess.run(['git', 'branch', '-M', 'main'],
                   capture_output=True, cwd=str(git_root))

    # 6. Asegurar al menos un commit (para que haya algo que pushear)
    r = subprocess.run(['git', 'log', '-1', '--oneline'],
                       capture_output=True, cwd=str(git_root))
    if r.returncode != 0:
        subprocess.run(['git', 'add', '.'], cwd=str(git_root))
        subprocess.run(['git', 'commit', '-m', 'Inicial: dashboard del Plan Estratégico'],
                       capture_output=True, cwd=str(git_root))
        print(f"   ✓ Commit inicial creado")

    # 7. Crear el repo Y hacer push (en un solo paso)
    print(f"   ⏳ Creando repo y subiendo archivos...")
    r = subprocess.run(
        ['gh', 'repo', 'create', repo_name,
         '--public', '--source=.', '--remote=origin', '--push'],
        capture_output=True, text=True, cwd=str(git_root),
    )
    if r.returncode != 0:
        err = (r.stderr or r.stdout).strip()
        if 'already exists' in err.lower() or 'name already exists' in err.lower():
            # El repo ya existe — solo conectar el remote y pushear
            print(f"   ℹ️  El repo {repo_name} ya existe. Conectando...")
            remote_url = f'https://github.com/{username}/{repo_name}.git'
            # Si origin ya existe, set-url; si no, add
            r_rem = subprocess.run(['git', 'remote'], capture_output=True, text=True, cwd=str(git_root))
            if 'origin' in r_rem.stdout.split():
                subprocess.run(['git', 'remote', 'set-url', 'origin', remote_url],
                               capture_output=True, cwd=str(git_root))
            else:
                subprocess.run(['git', 'remote', 'add', 'origin', remote_url],
                               capture_output=True, cwd=str(git_root))
            # Push
            r_push = subprocess.run(['git', 'push', '-u', 'origin', 'main'],
                                    capture_output=True, text=True, cwd=str(git_root))
            if r_push.returncode != 0:
                print(f"   ❌ Push falló: {(r_push.stderr or r_push.stdout).strip()}")
                return False
        else:
            print(f"   ❌ gh repo create falló:")
            for line in err.split('\n')[:6]:
                print(f"      {line}")
            return False
    print(f"   ✓ Repo creado y archivos subidos")

    # 8. Activar GitHub Pages (branch main, folder /docs)
    print(f"   ⏳ Activando GitHub Pages...")
    r = subprocess.run(
        ['gh', 'api', '--method', 'POST',
         f'/repos/{username}/{repo_name}/pages',
         '-f', 'source[branch]=main',
         '-f', 'source[path]=/docs'],
        capture_output=True, text=True,
    )
    if r.returncode == 0:
        print(f"   ✓ GitHub Pages activado (branch main, folder /docs)")
    else:
        err = (r.stderr or r.stdout).strip()
        if '409' in err or 'already' in err.lower():
            print(f"   ℹ️  GitHub Pages ya estaba activado")
        else:
            print(f"   ⚠️  No se pudo activar Pages automáticamente.")
            print(f"      Actívelo manualmente en:")
            print(f"      https://github.com/{username}/{repo_name}/settings/pages")

    pages_url = f"https://{username}.github.io/{repo_name}/"
    print(f"\n🎉 ¡Todo listo!")
    print(f"   📦 Repo:      https://github.com/{username}/{repo_name}")
    print(f"   🌐 Dashboard: {pages_url}")
    print(f"   ⏱️  La URL pública estará disponible en 1-2 minutos.")
    print(f"\n   Próximas veces ejecute solo:")
    print(f"      python \"Plan Estrategico Dashboard.py\"")
    return True


def publish_to_github(out_path):
    """Publica el dashboard a GitHub.

    Flujo automático:
      1. Si no hay repo Git → lo inicializa.
      2. Commitea el dashboard si hubo cambios.
      3. Si no hay remote configurado → intenta crear todo con gh CLI.
      4. Si hay remote → push normal.
    """
    import subprocess

    git_root = _find_git_root(out_path.parent)
    if git_root is None:
        # Auto-init
        try:
            subprocess.run(['git', 'init'], capture_output=True, cwd=str(Path.cwd()))
            subprocess.run(['git', 'branch', '-M', 'main'],
                           capture_output=True, cwd=str(Path.cwd()))
            git_root = Path.cwd()
            print(f"\n   ✓ Repositorio Git inicializado en {git_root.name}/")
        except Exception as e:
            print(f"\n   ⚠️  No se pudo inicializar Git: {e}")
            return False

    try:
        subprocess.run(['git', '--version'], capture_output=True, check=True, cwd=str(git_root))
    except (subprocess.CalledProcessError, FileNotFoundError):
        print(f"   ⚠️  Git no está instalado o no está en el PATH. Saltando publicación.")
        return False

    try:
        rel_path = str(out_path.resolve().relative_to(git_root))
    except ValueError:
        rel_path = str(out_path)

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    print(f"\n📤 Publicando en GitHub...")

    # Add + commit del dashboard (si cambió)
    r = subprocess.run(['git', 'add', rel_path],
                       capture_output=True, text=True, cwd=str(git_root))
    if r.returncode != 0:
        print(f"   ❌ git add falló: {r.stderr.strip() or r.stdout.strip()}")
        return False

    r = subprocess.run(['git', 'diff', '--cached', '--quiet'], cwd=str(git_root))
    has_pending_commit = (r.returncode != 0)
    if has_pending_commit:
        msg = f"Actualizar dashboard — {timestamp}"
        r = subprocess.run(['git', 'commit', '-m', msg],
                           capture_output=True, text=True, cwd=str(git_root))
        if r.returncode != 0:
            print(f"   ❌ git commit falló: {r.stderr.strip() or r.stdout.strip()}")
            return False
        print(f"   ✓ Commit creado: {msg}")

    # ¿Hay remote configurado?
    r = subprocess.run(['git', 'remote'], capture_output=True, text=True, cwd=str(git_root))
    has_remote = bool(r.stdout.strip())

    if not has_remote:
        # Primera vez: usar gh CLI para configurar todo
        return _try_gh_setup(git_root, out_path)

    # Push normal
    if not has_pending_commit:
        # Verificar si hay commits sin pushear
        r = subprocess.run(['git', 'status', '-sb'], capture_output=True, text=True, cwd=str(git_root))
        if 'ahead' not in r.stdout:
            print(f"   ℹ️  El dashboard no cambió desde el último push. Nada que publicar.")
            return True

    r = subprocess.run(['git', 'push'], capture_output=True, text=True, cwd=str(git_root))
    if r.returncode != 0:
        err = (r.stderr.strip() or r.stdout.strip())
        print(f"   ❌ git push falló:")
        for line in err.split('\n')[:6]:
            print(f"      {line}")
        return False
    print(f"   ✓ Dashboard publicado a GitHub")
    print(f"   🌐 Estará disponible en su URL de GitHub Pages en unos segundos.")
    return True


def _interactive_setup_github(git_root, out_path):
    """Pide token y nombre de repo interactivamente y llama a setup_github."""
    import getpass
    import os

    print("\n╔══════════════════════════════════════════════════════════════╗")
    print("║   CONFIGURACIÓN INICIAL DE GITHUB (sólo una vez)             ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()
    print("Necesita un Personal Access Token (PAT) de GitHub.")
    print()
    print("Si no tiene uno, créelo aquí (toma 30 segundos):")
    print("  → https://github.com/settings/tokens/new")
    print()
    print("  Configure así:")
    print("    Note:        Plan Estratégico Dashboard")
    print("    Expiration:  No expiration (o 90 días)")
    print("    Scopes:      marque la casilla 'repo' (todo el grupo)")
    print()
    print("  Clic 'Generate token' y copie el token (empieza con 'ghp_').")
    print()

    # Token: variable de entorno > prompt
    token = os.environ.get('GITHUB_TOKEN', '').strip()
    if token:
        print(f"✓ Usando token de la variable de entorno GITHUB_TOKEN")
    else:
        try:
            token = getpass.getpass("Pegue su Personal Access Token (no se mostrará): ").strip()
        except (KeyboardInterrupt, EOFError):
            print("\nCancelado.")
            return None
    if not token:
        print("❌ No se ingresó token. Cancelado.")
        return None

    # Nombre del repo: default = nombre de la carpeta sanitizado
    default_repo = git_root.name.lower()
    default_repo = ''.join(c if c.isalnum() or c in '-_' else '-' for c in default_repo)
    default_repo = default_repo.strip('-') or 'plan-estrategico'
    try:
        repo_name = input(f"Nombre del repo (Enter para usar '{default_repo}'): ").strip()
    except (KeyboardInterrupt, EOFError):
        print("\nCancelado.")
        return None
    if not repo_name:
        repo_name = default_repo

    return setup_github(token, repo_name, git_root, out_path)


# ============================================================
# MAIN
# ============================================================

def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    flags = [a for a in sys.argv[1:] if a.startswith('--')]
    no_publish = '--no-publish' in flags
    do_setup = '--setup-github' in flags

    xlsx_path = Path(args[0]) if args else Path('Tablero_Indicadores_PE.xlsx')
    out_path = Path(args[1]) if len(args) > 1 else Path('docs/index.html')

    if not xlsx_path.exists():
        print(f"❌ ERROR: archivo no encontrado: {xlsx_path}")
        sys.exit(1)

    print(f"📖 Leyendo  {xlsx_path}")
    indicators = process_workbook(xlsx_path)
    print(f"   ✓ {len(indicators)} indicadores procesados")

    total_warnings = sum(len(i['warnings']) for i in indicators)
    if total_warnings:
        print(f"   ⚠️  {total_warnings} avisos de validación (visibles en el dashboard)")

    payload = build_payload(indicators)
    print(f"   ✓ Resumen: {payload['summary']}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    html = render_dashboard_html(payload)
    out_path.write_text(html, encoding='utf-8')

    size_kb = out_path.stat().st_size / 1024
    print(f"✅ Dashboard generado: {out_path}  ({size_kb:.1f} KB)")

    # Setup inicial de GitHub si se solicita
    if do_setup:
        import subprocess
        git_root = _find_git_root(out_path.parent)
        if git_root is None:
            # Inicializar git en la carpeta actual
            print(f"\n🔧 Inicializando repositorio Git en {Path.cwd()}...")
            subprocess.run(['git', 'init'], capture_output=True, cwd=str(Path.cwd()))
            subprocess.run(['git', 'branch', '-M', 'main'], capture_output=True, cwd=str(Path.cwd()))
            git_root = Path.cwd()
            # Commit inicial con todo
            subprocess.run(['git', 'add', '.'], cwd=str(git_root))
            subprocess.run(['git', 'commit', '-m', 'Inicial'], capture_output=True, cwd=str(git_root))
        _interactive_setup_github(git_root, out_path)
        return

    # Auto-publicar a GitHub (a menos que se use --no-publish)
    if not no_publish:
        publish_to_github(out_path)
    else:
        print(f"\n   (Auto-publicación desactivada con --no-publish)")



if __name__ == '__main__':
    main()
